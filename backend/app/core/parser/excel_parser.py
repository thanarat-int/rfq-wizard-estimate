from openpyxl import load_workbook
import io


def extract_text_from_excel(file_content: bytes) -> str:
    """Extract text content from an Excel file."""
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip(" |"):
                text_parts.append(row_text)

    wb.close()
    return "\n".join(text_parts)


def extract_data_from_excel(file_content: bytes) -> list[dict]:
    """Extract structured data from Excel (for master data import)."""
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    all_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First row as headers
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        for row in rows[1:]:
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = cell
            if any(v is not None for v in row_dict.values()):
                all_data.append(row_dict)

    wb.close()
    return all_data
