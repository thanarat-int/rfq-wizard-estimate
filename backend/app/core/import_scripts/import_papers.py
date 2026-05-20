"""Import paper prices from Paper Price (MI2).xlsx into DB."""
import os
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.models.master_data import Paper


DATA_FILE = os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", "data", "Paper Price (MI2).xlsx")


def run_import(db: Session) -> dict:
    """Read xlsx and upsert all papers into DB."""
    filepath = os.path.normpath(DATA_FILE)
    if not os.path.exists(filepath):
        return {"success": False, "imported": 0, "errors": [f"File not found: {filepath}"]}

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    imported = 0
    updated = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            if not row or not row[1]:  # skip empty rows, need paper_code
                continue

            xlsx_id = row[0]
            paper_code = str(row[1]).strip() if row[1] else None
            paper_type = str(row[2]).strip() if row[2] else None
            gram = _int(row[3])
            thickness_micron = _float(row[4])
            thickness_mm = _float(row[5])
            price = _float(row[6])
            special_ink = str(row[7]).strip() if row[7] else None
            category_id = _int(row[8])
            is_fsc = bool(_int(row[9]))
            project_code = str(row[10]).strip() if row[10] else None
            price_import = _float(row[11])
            is_only_baht_per_sheet = _int(row[12]) or 0
            only_print_type = str(row[13]).strip() if row[13] and str(row[13]).strip().upper() != "NULL" else None
            only_print_type_id = _int(row[14])
            brand = str(row[15]).strip() if row[15] else None
            supplier = str(row[16]).strip() if row[16] else None
            brand_import = str(row[17]).strip() if len(row) > 17 and row[17] else None
            supplier_import = str(row[18]).strip() if len(row) > 18 and row[18] else None

            # Derive legacy fields
            name = f"{paper_code} {gram}g {paper_type}" if paper_code and gram else paper_type
            price_per_kg = price if is_only_baht_per_sheet == 0 else None
            price_per_sheet = price if is_only_baht_per_sheet == 1 else None

            # Upsert by paper_code + gram
            existing = db.query(Paper).filter(
                Paper.paper_code == paper_code,
                Paper.gram == gram
            ).first()

            data = dict(
                paper_code=paper_code,
                paper_type=paper_type,
                gram=gram,
                thickness_micron=thickness_micron,
                thickness_mm=thickness_mm,
                price=price,
                special_ink_paper_code=special_ink,
                category_id=category_id,
                is_fsc=is_fsc,
                project_code=project_code,
                price_import=price_import,
                is_only_baht_per_sheet=is_only_baht_per_sheet,
                only_print_type=only_print_type,
                only_print_type_id=only_print_type_id,
                brand=brand,
                supplier=supplier,
                brand_import=brand_import,
                supplier_import=supplier_import,
                # Legacy
                name=name,
                type=paper_type,
                gsm=gram,
                price_per_kg=price_per_kg,
                price_per_sheet=price_per_sheet,
                active=True,
            )

            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                db.add(Paper(**data))
                imported += 1

        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()
    return {
        "success": True,
        "imported": imported,
        "updated": updated,
        "total": imported + updated,
        "errors": errors,
    }


def _int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None
